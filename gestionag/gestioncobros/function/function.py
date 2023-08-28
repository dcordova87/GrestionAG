from openpyxl import load_workbook
from openpyxl.worksheet.pagebreak import Break
from gestionag.settings import MEDIA_ROOT as mediadir
import os


def GenerarListado(listado, pendiente):
    #variables de rutas y hojas
    #listado_xlsx = 'listado.xlsx'
    #pendiente_xlsx = 'pendiente.xlsx'
    sheetNameListado = 'CUP'
    sheetNamePendiente = 'Sheet1'

    listado = load_workbook(listado) #abrir excel listado
    pendiente = load_workbook(pendiente) #abrir excel pendiente

    #sheetsL = listado.sheetnames  #todas las hojas del libro Listado
    #sheetsP = pendiente.sheetnames #todas las hojas del libro Pendiente

    sheetL = listado[sheetNameListado] # Seleccionar hoja del listado de cobros
    sheetP = pendiente[sheetNamePendiente] # Seleccionar hoja del excel con el pendiente
   

    for rowNumL in range(2, sheetL.max_row+1): #recorrer hoja listado

        contador = 0 
        #print(sheetL.cell(row=rowNumL, column=3).value)

        for rowNumP in range(2, sheetP.max_row+1): # recorrer hoja pendiente
            if contador == 1:
                break
            #print(sheetP.cell(row=rowNumP, column=2).value)

            if str(sheetL.cell(row=rowNumL, column=3).value) == str(sheetP.cell(row=rowNumP, column=2).value):
                sheetL.cell(row=rowNumL, column=5).value = sheetP.cell(row=rowNumP, column=3).value
                contador = 1
                #break
                
        if contador == 0:

            sheetL.cell(row=rowNumL, column=5).value = 0

    ruta_guardado = os.path.join(mediadir, 'listado1.xlsx')
    #ruta_guardado = 'gestioncobros/download/listado1.xlsx'
    listado.save(ruta_guardado)
    return ruta_guardado